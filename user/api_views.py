from rest_framework import status, viewsets, filters
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.pagination import PageNumberPagination
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import timedelta, date
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.db import models

from .models import Customer, Center, Slot, SlotBooking, Plan, Invoice, User
from .serializers import (
    CustomerSerializer, LoginSerializer, UserSerializer,
    CenterSerializer, SlotSerializer, SlotBookingSerializer,
    PlanSerializer, InvoiceSerializer,
)


# =========================================
# HELPERS
# =========================================

def custom_response(success=True, message="", data=None, status_code=200):
    return Response({"success": success, "message": message, "data": data}, status=status_code)


class StandardPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

    def get_paginated_response(self, data, message=""):
        return Response({
            "success": True,
            "message": message,
            "data": data,
            "pagination": {
                "count": self.page.paginator.count,
                "total_pages": self.page.paginator.num_pages,
                "current_page": self.page.number,
                "next": self.get_next_link(),
                "previous": self.get_previous_link(),
            }
        })


# =========================================
# LOGIN API
# =========================================

class LoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        is_admin = request.query_params.get('is_admin') == 'true'
        is_branch = request.query_params.get('is_branch') == 'true'

        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data["user"]

            if is_admin and user.role != 'super_admin':
                return custom_response(False, "Access denied. Not an admin.", None, status.HTTP_403_FORBIDDEN)
            if is_branch and user.role != 'branch_user':
                return custom_response(False, "Access denied. Not a branch user.", None, status.HTTP_403_FORBIDDEN)

            refresh = RefreshToken.for_user(user)
            return custom_response(True, "Login successful", {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": user.role,
                    "center": user.center.center_name if user.center else None,
                    "center_id": user.center.id if user.center else None,
                },
                "redirect_to": "admin_dashboard" if user.role == "super_admin" else "branch_dashboard",
            })
        return custom_response(False, "Invalid credentials", serializer.errors, status.HTTP_400_BAD_REQUEST)


# =========================================
# USER VIEWSET
# =========================================

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-id')
    serializer_class = UserSerializer
    pagination_class = StandardPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['username', 'email', 'first_name', 'last_name']

    def get_queryset(self):
        qs = super().get_queryset()
        role = self.request.query_params.get('role')
        if role:
            qs = qs.filter(role=role)
        return qs

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Users fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Users fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_object())
        return custom_response(True, "User fetched successfully", serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "User created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "User updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "User deleted successfully")

    @action(detail=False, methods=['get'], url_path='by-role/(?P<role>[^/.]+)')
    def by_role(self, request, role=None):
        users = self.get_queryset().filter(role=role)
        serializer = self.get_serializer(users, many=True)
        return custom_response(True, f"Users with role '{role}' fetched successfully", serializer.data)

    @action(detail=True, methods=['post'], url_path='change-password')
    def change_password(self, request, pk=None):
        user = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password:
            return custom_response(False, "new_password is required", None, status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return custom_response(False, "New password must be at least 6 characters", None, status.HTTP_400_BAD_REQUEST)
        
        user.set_password(new_password)
        user.save()
        return custom_response(True, "Password changed successfully", None)


# =========================================
# CENTER VIEWSET
# =========================================

class CenterViewSet(viewsets.ModelViewSet):
    queryset = Center.objects.all().order_by('-id')
    serializer_class = CenterSerializer
    pagination_class = StandardPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['center_name', 'location', 'mobile']

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Centers fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Centers fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Center fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Center created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Center updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "Center deleted successfully")

    @action(detail=True, methods=['post'], url_path='change-password')
    def change_password(self, request, pk=None):
        center = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password:
            return custom_response(False, "new_password is required", None, status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return custom_response(False, "New password must be at least 6 characters", None, status.HTTP_400_BAD_REQUEST)
        
        # Find the branch user associated with this center
        branch_user = User.objects.filter(center=center, role='branch_user').first()
        
        if not branch_user:
            return custom_response(False, "No branch user found for this center", None, status.HTTP_404_NOT_FOUND)
        
        # Update the password
        branch_user.set_password(new_password)
        branch_user.save()
        
        return custom_response(True, f"Password changed successfully for {center.center_name}", {
            "center_id": center.id,
            "center_name": center.center_name,
            "user_email": branch_user.email
        })


class CenterMinimalView(APIView):
    def get(self, request):
        from .serializers import CenterMinimalSerializer
        centers = Center.objects.all().order_by('center_name')
        serializer = CenterMinimalSerializer(centers, many=True)
        return custom_response(True, "Centers fetched successfully", serializer.data)


# =========================================
# PLAN VIEWSET
# =========================================

class PlanViewSet(viewsets.ModelViewSet):
    queryset = Plan.objects.all().order_by('-id')
    serializer_class = PlanSerializer
    pagination_class = StandardPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['plan_name']

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Plans fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Plans fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Plan fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Plan created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Plan updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        plan = self.get_object()
        if Invoice.objects.filter(plan=plan).exists():
            return custom_response(False, "Cannot delete plan. It is used in invoices.", None, status.HTTP_400_BAD_REQUEST)
        plan.delete()
        return custom_response(True, "Plan deleted successfully")


# =========================================
# CUSTOMER VIEWSET
# =========================================

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by('-id')
    serializer_class = CustomerSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search')
        center = self.request.query_params.get('center')
        date_param = self.request.query_params.get('date')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(mobile__icontains=search))
        if center:
            qs = qs.filter(center_id=center)
        if date_param:
            qs = qs.filter(created_at__date=date_param)
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)
        return qs

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Customers fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Customers fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Customer fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        if not data.get('center_id') and hasattr(request.user, 'center') and request.user.center:
            data['center_id'] = request.user.center.id
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Customer created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Customer updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "Customer deleted successfully")

    @action(detail=False, methods=['get'], url_path='wave-choices')
    def wave_choices(self, request):
        choices = [{'value': k, 'label': v} for k, v in Customer.WAVE_CHOICES]
        return custom_response(True, "Wave choices fetched successfully", choices)

    @action(detail=False, methods=['get'], url_path='minimal')
    def minimal(self, request):
        center = request.query_params.get('center')
        qs = Customer.objects.only('id', 'name').order_by('-id')
        if center:
            qs = qs.filter(center_id=center)
        data = [{'id': c.id, 'name': c.name} for c in qs]
        return custom_response(True, "Customers fetched successfully", data)


# =========================================
# SLOT VIEWSET
# =========================================

class SlotViewSet(viewsets.ModelViewSet):
    queryset = Slot.objects.all().order_by('-id')
    serializer_class = SlotSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        return super().get_queryset()

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Slots fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Slots fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Slot fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Slot created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Slot updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "Slot deleted successfully")

    @action(detail=True, methods=['get'], url_path='customers')
    def customers(self, request, pk=None):
        slot = self.get_object()
        booking_date = request.query_params.get('date')
        bookings = SlotBooking.objects.filter(slot=slot).select_related('customer')
        if booking_date:
            bookings = bookings.filter(booking_date=booking_date)
        data = {
            'slot': {
                'id': slot.id,
                'start_time': slot.start_time.strftime('%I:%M %p'),
                'end_time': slot.end_time.strftime('%I:%M %p'),
                'total_slot': slot.total_slot,
            },
            'bookings': [
                {
                    'booking_id': b.id,
                    'booking_date': str(b.booking_date),
                    'status': b.status,
                    'customer': {
                        'id': b.customer.id,
                        'name': b.customer.name,
                        'mobile': b.customer.mobile,
                        'email': b.customer.email,
                    }
                }
                for b in bookings
            ]
        }
        return custom_response(True, "Customers for slot fetched successfully", data)

    @action(detail=False, methods=['get'], url_path='minimal')
    def minimal(self, request):
        booking_date = request.query_params.get('date')
        center_id = request.query_params.get('center')
        
        if not booking_date:
            return custom_response(False, "date parameter is required", None, status.HTTP_400_BAD_REQUEST)
        
        # Get all enabled slots
        slots = Slot.objects.filter(is_enabled=True).order_by('start_time')
        
        available_slots = []
        for slot in slots:
            # Count bookings for this slot on the given date (filtered by center if provided)
            booking_filter = {'slot': slot, 'booking_date': booking_date}
            if center_id:
                booking_filter['center_id'] = center_id
            
            booked_count = SlotBooking.objects.filter(**booking_filter).count()
            
            # Calculate balance (remaining capacity)
            balance = slot.total_slot - booked_count
            
            # Only include slots that have available capacity
            if balance > 0:
                available_slots.append({
                    'id': slot.id,
                    'start_time': slot.start_time.strftime('%I:%M %p'),
                    'end_time': slot.end_time.strftime('%I:%M %p'),
                    'total_slot': slot.total_slot,
                    'booked_slot': booked_count,
                    'balance_slot': balance,
                })
        
        return custom_response(True, "Available slots fetched successfully", available_slots)


# =========================================
# SLOT BOOKING VIEWSET
# =========================================

class SlotBookingViewSet(viewsets.ModelViewSet):
    queryset = SlotBooking.objects.all().order_by('-id')
    serializer_class = SlotBookingSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        date_param = self.request.query_params.get('date')
        center_param = self.request.query_params.get('center')
        
        if date_param:
            qs = qs.filter(booking_date=date_param)
        if center_param:
            qs = qs.filter(center_id=center_param)
        
        return qs

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Slot bookings fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Slot bookings fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Slot booking fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Slot booking created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Slot booking updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "Slot booking deleted successfully")


# =========================================
# INVOICE VIEWSET
# =========================================

class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all().order_by('-id')
    serializer_class = InvoiceSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search')
        center = self.request.query_params.get('center')
        date_param = self.request.query_params.get('date')
        plan = self.request.query_params.get('plan')
        if search:
            qs = qs.filter(Q(customer__name__icontains=search))
        if center:
            qs = qs.filter(center_id=center)
        if date_param:
            qs = qs.filter(date=date_param)
        if plan:
            qs = qs.filter(plan_id=plan)
        return qs

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.paginator.get_paginated_response(serializer.data, "Invoices fetched successfully")
        serializer = self.get_serializer(qs, many=True)
        return custom_response(True, "Invoices fetched successfully", serializer.data)

    def retrieve(self, request, *args, **kwargs):
        return custom_response(True, "Invoice fetched successfully", self.get_serializer(self.get_object()).data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Invoice created successfully", serializer.data, status.HTTP_201_CREATED)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(self.get_object(), data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return custom_response(True, "Invoice updated successfully", serializer.data)
        return custom_response(False, "Validation error", serializer.errors, status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        self.get_object().delete()
        return custom_response(True, "Invoice deleted successfully")


# =========================================
# ADMIN DASHBOARD API
# =========================================

class AdminDashboardView(APIView):

    def get(self, request):
        today = date.today()
        this_month_start = today.replace(day=1)
        last_month_end = this_month_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        total_customers = Customer.objects.count()
        total_revenue = Invoice.objects.aggregate(total=Sum('amount'))['total'] or 0
        total_sessions = Slot.objects.count()
        
        # Calculate overall booking rate based on slot capacity
        total_capacity = Slot.objects.aggregate(total=Sum('total_slot'))['total'] or 0
        total_booked = SlotBooking.objects.count()
        booking_rate = round((total_booked / total_capacity * 100), 2) if total_capacity > 0 else 0

        customers_this_month = Customer.objects.filter(created_at__date__gte=this_month_start).count()
        customers_last_month = Customer.objects.filter(created_at__date__gte=last_month_start, created_at__date__lte=last_month_end).count()
        revenue_this_month = Invoice.objects.filter(date__gte=this_month_start).aggregate(total=Sum('amount'))['total'] or 0
        revenue_last_month = Invoice.objects.filter(date__gte=last_month_start, date__lte=last_month_end).aggregate(total=Sum('amount'))['total'] or 0
        sessions_this_month = Slot.objects.filter(created_at__date__gte=this_month_start).count()
        sessions_last_month = Slot.objects.filter(created_at__date__gte=last_month_start, created_at__date__lte=last_month_end).count()
        
        # Booking rate calculations for growth
        booked_this_month = SlotBooking.objects.filter(booking_date__gte=this_month_start).count()
        booked_last_month = SlotBooking.objects.filter(booking_date__gte=last_month_start, booking_date__lte=last_month_end).count()
        
        booking_rate_this_month = round((booked_this_month / total_capacity * 100), 2) if total_capacity > 0 else 0
        booking_rate_last_month = round((booked_last_month / total_capacity * 100), 2) if total_capacity > 0 else 0

        def growth(cur, prev):
            if prev == 0:
                return 100.0 if cur > 0 else 0.0
            return round((cur - prev) / prev * 100, 2)

        # Center-wise performance
        centers_data = []
        # Get total slot capacity (shared across all centers)
        total_capacity = Slot.objects.aggregate(total=Sum('total_slot'))['total'] or 0
        
        for center in Center.objects.all():
            c_customers = Customer.objects.filter(center=center).count()
            c_revenue = Invoice.objects.filter(center=center).aggregate(total=Sum('amount'))['total'] or 0
            
            # Get total bookings for this center (all time)
            c_total_booked = SlotBooking.objects.filter(center=center).count()
            
            # Calculate booking rate: (Center's bookings / Total slot capacity) × 100
            c_rate = round((c_total_booked / total_capacity * 100), 2) if total_capacity > 0 else 0
            
            centers_data.append({
                'center_id': center.id,
                'center_name': center.center_name,
                'customers': c_customers,
                'revenue': float(c_revenue),
                'booking_rate': c_rate,
                'total_bookings': c_total_booked,
            })

        # Revenue last 7 months
        revenue_labels, revenue_values = [], []
        for i in range(6, -1, -1):
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            rev = Invoice.objects.filter(date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or 0
            revenue_labels.append(date(year, month, 1).strftime('%b'))
            revenue_values.append(float(rev))

        # Membership by plan
        membership_data = [
            {'plan_name': p.plan_name, 'customer_count': Customer.objects.filter(plan=p).count()}
            for p in Plan.objects.all()
        ]

        # Recent customers (limit to 5)
        recent_customers = [
            {
                "id": c.id, "name": c.name, "mobile": c.mobile,
                "center": c.center.center_name if c.center else None,
                "plan": c.plan.plan_name if c.plan else None,
                "status": c.status, "joined": str(c.created_at.date()),
            }
            for c in Customer.objects.select_related("center", "plan").order_by("-created_at")[:5]
        ]

        # Slot bookings last 7 months
        booking_labels, booked_values, free_values = [], [], []
        # Get current total slot capacity (this is constant)
        total_slot_capacity = Slot.objects.aggregate(total=Sum('total_slot'))['total'] or 0
        
        for i in range(6, -1, -1):
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            
            # Get bookings for this month
            month_bookings = SlotBooking.objects.filter(
                booking_date__year=year, 
                booking_date__month=month
            ).count()
            
            # Calculate free slots (total capacity - bookings for this month)
            free = max(total_slot_capacity - month_bookings, 0)
            
            booking_labels.append(date(year, month, 1).strftime('%b'))
            booked_values.append(month_bookings)
            free_values.append(free)

        page = int(request.query_params.get("page", 1))
        page_size = 10

        def paginate(data):
            start = (page - 1) * page_size
            total = len(data)
            total_pages = max(1, (total + page_size - 1) // page_size)
            has_next = page < total_pages
            has_previous = page > 1
            return {
                "results": data[start:start + page_size],
                "count": total,
                "total_pages": total_pages,
                "current_page": page,
                "next": page + 1 if has_next else None,
                "previous": page - 1 if has_previous else None,
            }

        return custom_response(True, "Admin dashboard fetched successfully", {
            "summary": {
                "total_customers": total_customers,
                "total_revenue": float(total_revenue),
                "total_sessions": total_sessions,
                "booking_rate": booking_rate,
                "customer_growth": growth(customers_this_month, customers_last_month),
                "revenue_growth": growth(float(revenue_this_month), float(revenue_last_month)),
                "session_growth": growth(sessions_this_month, sessions_last_month),
                "booking_rate_growth": growth(booking_rate_this_month, booking_rate_last_month),
            },
            "centerwise_performance": paginate(centers_data),
            "revenue_overview": {"labels": revenue_labels, "values": revenue_values},
            "slot_bookings_overview": {
                "labels": booking_labels, 
                "booked": booked_values, 
                "free": free_values
            },
            "membership_status": paginate(membership_data),
            "recent_customers": recent_customers,
        })


# =========================================
# SLOT BOOKINGS DASHBOARD API
# =========================================

class SlotBookingsDashboardView(APIView):

    def get(self, request):
        from django.db.models.functions import ExtractYear, ExtractMonth

        today = date.today()
        months = []
        for i in range(6, -1, -1):
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            months.append((year, month))

        slot_counts = (
            Slot.objects.annotate(year=ExtractYear('created_at'), month=ExtractMonth('created_at'))
            .values('year', 'month')
            .annotate(total=Count('id'))
        )
        slot_map = {(s['year'], s['month']): s['total'] for s in slot_counts}

        booking_counts = (
            SlotBooking.objects.filter(status__iexact='booked')
            .annotate(year=ExtractYear('booking_date'), month=ExtractMonth('booking_date'))
            .values('year', 'month')
            .annotate(total=Count('id'))
        )
        booking_map = {(b['year'], b['month']): b['total'] for b in booking_counts}

        data, labels, booked_list, free_list = [], [], [], []

        for year, month in months:
            label = date(year, month, 1).strftime('%b')
            total_slots = slot_map.get((year, month), 0)
            booked = booking_map.get((year, month), 0)
            free = max(total_slots - booked, 0)

            data.append({'month': label, 'year': year, 'booked': booked, 'free': free})
            labels.append(label)
            booked_list.append(booked)
            free_list.append(free)

        return custom_response(True, "Slot bookings overview fetched successfully", {
            'data': data,
            'labels': labels,
            'booked': booked_list,
            'free': free_list,
        })


# =========================================
# CUSTOMER EXCEL DOWNLOAD
# =========================================

class CustomerExcelDownloadView(APIView):

    def get(self, request, pk=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Customers'

        headers = ['ID', 'Name', 'Mobile', 'Email', 'Center', 'Plan', 'Wave', 'Start Date', 'Expiry Date', 'Status']
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        if pk:
            try:
                customers = [Customer.objects.select_related('center', 'plan').get(pk=pk)]
            except Customer.DoesNotExist:
                return custom_response(False, 'Customer not found', None, status.HTTP_404_NOT_FOUND)
            filename = f'customer_{pk}.xlsx'
        else:
            center = request.query_params.get('center')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            qs = Customer.objects.select_related('center', 'plan').all().order_by('-id')
            if center:
                qs = qs.filter(center_id=center)
            if start_date:
                qs = qs.filter(start_date__gte=start_date)
            if end_date:
                qs = qs.filter(start_date__lte=end_date)
            customers = qs
            filename = 'customers.xlsx'

        for row, c in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=c.name)
            ws.cell(row=row, column=3, value=c.mobile)
            ws.cell(row=row, column=4, value=c.email or '')
            ws.cell(row=row, column=5, value=c.center.center_name if c.center else '')
            ws.cell(row=row, column=6, value=c.plan.plan_name if c.plan else '')
            ws.cell(row=row, column=7, value=c.wave)
            ws.cell(row=row, column=8, value=str(c.start_date) if c.start_date else '')
            ws.cell(row=row, column=9, value=str(c.expiry_date) if c.expiry_date else '')
            ws.cell(row=row, column=10, value=c.status)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


# =========================================
# INVOICE EXCEL DOWNLOAD
# =========================================

class InvoiceExcelDownloadView(APIView):

    def get(self, request, pk=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Invoices'

        headers = ['Invoice ID', 'Customer', 'Center', 'Plan', 'Amount', 'Date', 'Status']
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        if pk:
            try:
                invoices = [Invoice.objects.select_related('customer', 'center', 'plan').get(pk=pk)]
            except Invoice.DoesNotExist:
                return custom_response(False, 'Invoice not found', None, status.HTTP_404_NOT_FOUND)
            filename = f'invoice_{pk}.xlsx'
        else:
            invoices = Invoice.objects.select_related('customer', 'center', 'plan').all().order_by('-id')
            filename = 'invoices.xlsx'

        for row, inv in enumerate(invoices, 2):
            ws.cell(row=row, column=1, value=f'INV-{inv.id:03d}')
            ws.cell(row=row, column=2, value=inv.customer.name if inv.customer else '')
            ws.cell(row=row, column=3, value=inv.center.center_name if inv.center else '')
            ws.cell(row=row, column=4, value=inv.plan.plan_name if inv.plan else '')
            ws.cell(row=row, column=5, value=float(inv.amount))
            ws.cell(row=row, column=6, value=str(inv.date))
            ws.cell(row=row, column=7, value=inv.status)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


# =========================================
# CHANGE PASSWORD API
# =========================================

class ChangePasswordView(APIView):

    def post(self, request):
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not all([current_password, new_password, confirm_password]):
            return custom_response(False, "All fields are required.", None, status.HTTP_400_BAD_REQUEST)

        if not request.user.check_password(current_password):
            return custom_response(False, "Current password is incorrect.", None, status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return custom_response(False, "New password and confirm password do not match.", None, status.HTTP_400_BAD_REQUEST)

        request.user.set_password(new_password)
        request.user.save()
        return custom_response(True, "Password updated successfully.")


# =========================================
# BRANCH SETTINGS API
# =========================================

class BranchChangePasswordView(APIView):

    def post(self, request):
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not all([old_password, new_password, confirm_password]):
            return custom_response(False, "All fields are required.", None, status.HTTP_400_BAD_REQUEST)

        if not request.user.check_password(old_password):
            return custom_response(False, "Old password is incorrect.", None, status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return custom_response(False, "New password and confirm password do not match.", None, status.HTTP_400_BAD_REQUEST)

        request.user.set_password(new_password)
        request.user.save()
        return custom_response(True, "Password updated successfully.")


# =========================================
# BRANCH DASHBOARD API
# =========================================

class BranchDashboardView(APIView):

    def get(self, request):
        today = timezone.now().date()
        center_id = request.query_params.get('center_id') or request.query_params.get('center')
        if not center_id and hasattr(request.user, 'center') and request.user.center:
            center_id = request.user.center.id

        if not center_id:
            return custom_response(False, "center_id is required. Pass ?center_id=1 or login as a branch user.", None, status.HTTP_400_BAD_REQUEST)

        try:
            center = Center.objects.get(pk=center_id)
        except Center.DoesNotExist:
            return custom_response(False, "Center not found", None, status.HTTP_404_NOT_FOUND)

        # Get all slots
        slots = Slot.objects.all()
        total_slots = slots.aggregate(total=Sum('total_slot'))['total'] or 0
        
        # Filter bookings by center for today
        booked_today = SlotBooking.objects.filter(booking_date=today, center_id=center_id).count()
        free_slots = total_slots - booked_today
        booking_rate = round((booked_today / total_slots * 100), 1) if total_slots > 0 else 0

        # Most purchased plan for this center
        most_purchased = (
            Invoice.objects.filter(center=center)
            .values('plan__plan_name')
            .annotate(count=Count('id'))
            .order_by('-count')
            .first()
        )

        # Today's slots with bookings for this center only
        today_slots = []
        for slot in slots.order_by('start_time'):
            booking = SlotBooking.objects.filter(
                slot=slot, 
                booking_date=today, 
                center_id=center_id
            ).select_related('customer').first()
            today_slots.append({
                "id": slot.id,
                "start_time": slot.start_time.strftime('%I:%M %p'),
                "end_time": slot.end_time.strftime('%I:%M %p'),
                "status": "booked" if booking else "available",
                "customer_name": booking.customer.name if booking else None,
            })

        # Recent customers for this center
        recent_customers_data = []
        for c in Customer.objects.filter(center=center).order_by('-created_at')[:5]:
            diff = (today - c.created_at.date()).days
            joined = "Today" if diff == 0 else "Yesterday" if diff == 1 else c.created_at.strftime('%d %b %Y')
            recent_customers_data.append({
                "id": c.id,
                "name": c.name,
                "plan": c.plan.plan_name if c.plan else None,
                "joined": joined,
            })

        return custom_response(True, "Branch dashboard fetched successfully", {
            "center_name": center.center_name,
            "center_email": center.email,
            "stats": {
                "slots_booked_today": booked_today,
                "total_slots": total_slots,
                "free_slots": free_slots,
                "booking_rate": f"{booking_rate}%",
                "most_purchased_plan": most_purchased['plan__plan_name'] if most_purchased else None,
            },
            "today_slots": today_slots,
            "recent_customers": recent_customers_data,
        })

# =========================================
# REPORTS API
# =========================================

class CustomerReportView(APIView):

    def get(self, request):
        search = request.query_params.get('search')
        center = request.query_params.get('center')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'

        qs = Customer.objects.select_related('center', 'plan').order_by('-created_at')

        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(mobile__icontains=search) | Q(email__icontains=search))
        if center:
            qs = qs.filter(center_id=center)
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        if export:
            return self._export_excel(qs)

        data = [
            {
                "id": c.id,
                "name": c.name,
                "mobile": c.mobile,
                "email": c.email,
                "center": c.center.center_name if c.center else None,
                "plan": c.plan.plan_name if c.plan else None,
                "start_date": str(c.start_date),
                "expiry_date": str(c.expiry_date),
                "status": c.status,
                "joined": str(c.created_at.date()),
            }
            for c in qs
        ]
        return custom_response(True, "Customer report fetched successfully", data)

    def _export_excel(self, qs):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        headers = ["ID", "Name", "Mobile", "Email", "Center", "Plan", "Start Date", "Expiry Date", "Status", "Joined"]
        ws.append(headers)

        for c in qs:
            ws.append([
                c.id, c.name, c.mobile, c.email or "",
                c.center.center_name if c.center else "",
                c.plan.plan_name if c.plan else "",
                str(c.start_date), str(c.expiry_date),
                c.status, str(c.created_at.date()),
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_report.xlsx"'
        wb.save(response)
        return response


class SlotBookingReportView(APIView):

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'

        booking_qs = SlotBooking.objects.select_related('slot', 'customer')
        if start_date:
            booking_qs = booking_qs.filter(booking_date__gte=start_date)
        if end_date:
            booking_qs = booking_qs.filter(booking_date__lte=end_date)

        slot_data = []
        for slot in Slot.objects.all():
            total_booked = booking_qs.filter(slot=slot).count()
            utilization = round((total_booked / slot.total_slot * 100), 1) if slot.total_slot > 0 else 0
            if utilization == 100:
                util_status = 'full'
            elif utilization >= 90:
                util_status = 'high'
            elif utilization >= 70:
                util_status = 'medium'
            else:
                util_status = 'low'
            slot_data.append({
                'slot_id': slot.id,
                'slot_time': f"{slot.start_time.strftime('%I:%M %p')} - {slot.end_time.strftime('%I:%M %p')}",
                'total_booked': total_booked,
                'total_capacity': slot.total_slot,
                'utilization': f"{utilization}%",
                'status': util_status,
            })

        if export:
            return self._export_excel(slot_data)
        return custom_response(True, "Slot booking report fetched successfully", slot_data)

    def _export_excel(self, slot_data):
        import openpyxl
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Slot Booking Report"
        ws.append(['Slot', 'Total Booked', 'Total Capacity', 'Utilization', 'Status'])
        for row in slot_data:
            ws.append([row['slot_time'], row['total_booked'], row['total_capacity'], row['utilization'], row['status']])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="slot_booking_report.xlsx"'
        wb.save(response)
        return response







# ============ BRANCH REPORTS ============

class BranchCustomerReportView(APIView):
    """Get customers for a specific branch"""
    
    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        if not center_id:
            return custom_response(False, "center_id is required", None, status.HTTP_400_BAD_REQUEST)
        
        try:
            center = Center.objects.get(pk=center_id)
        except Center.DoesNotExist:
            return custom_response(False, "Center not found", None, status.HTTP_404_NOT_FOUND)
        
        # Get customers for this branch
        customers_qs = Customer.objects.filter(center_id=center_id).select_related('plan').order_by('-created_at')
        if start_date:
            customers_qs = customers_qs.filter(created_at__date__gte=start_date)
        if end_date:
            customers_qs = customers_qs.filter(created_at__date__lte=end_date)
        
        customers_data = [
            {
                "id": c.id,
                "name": c.name,
                "mobile": c.mobile,
                "plan": c.plan.plan_name if c.plan else None,
                "joined": str(c.created_at.date()),
                "status": c.status,
            }
            for c in customers_qs
        ]
        
        # Export to Excel
        if export:
            return self._export_excel(center.center_name, customers_data)
        
        # Pagination
        total = len(customers_data)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = customers_data[start:end]
        
        return Response({
            "success": True,
            "message": f"Customers for {center.center_name}",
            "data": paginated,
            "pagination": {
                "count": total,
                "total_pages": max(1, (total + page_size - 1) // page_size),
                "current_page": page,
                "page_size": page_size,
            }
        })
    
    def _export_excel(self, center_name, customers_data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        headers = ['ID', 'Name', 'Mobile', 'Plan', 'Joined', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, c in enumerate(customers_data, 2):
            ws.cell(row=row, column=1, value=c['id'])
            ws.cell(row=row, column=2, value=c['name'])
            ws.cell(row=row, column=3, value=c['mobile'])
            ws.cell(row=row, column=4, value=c['plan'] or '')
            ws.cell(row=row, column=5, value=c['joined'])
            ws.cell(row=row, column=6, value=c['status'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="branch_customers_{center_name}.xlsx"'
        wb.save(response)
        return response


class BranchSlotBookingReportView(APIView):
    """Get slot bookings for a specific branch"""
    
    def get(self, request):
        center_id = request.query_params.get('center_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        if not center_id:
            return custom_response(False, "center_id is required", None, status.HTTP_400_BAD_REQUEST)
        
        try:
            center = Center.objects.get(pk=center_id)
        except Center.DoesNotExist:
            return custom_response(False, "Center not found", None, status.HTTP_404_NOT_FOUND)
        
        # Get slot bookings for this branch
        bookings_qs = SlotBooking.objects.filter(center_id=center_id).select_related('slot', 'customer').order_by('-booking_date')
        if start_date:
            bookings_qs = bookings_qs.filter(booking_date__gte=start_date)
        if end_date:
            bookings_qs = bookings_qs.filter(booking_date__lte=end_date)
        
        # Group slot bookings by slot
        slot_bookings_data = []
        slots = Slot.objects.all().order_by('start_time')
        for slot in slots:
            slot_bookings = bookings_qs.filter(slot=slot)
            total_booked = slot_bookings.count()
            
            if total_booked > 0:
                utilization = round((total_booked / slot.total_slot * 100), 1) if slot.total_slot > 0 else 0
                if utilization >= 90:
                    util_status = "high"
                elif utilization >= 70:
                    util_status = "medium"
                else:
                    util_status = "low"
                
                slot_bookings_data.append({
                    "slot": f"{slot.start_time.strftime('%I:%M %p')} - {slot.end_time.strftime('%I:%M %p')}",
                    "total_booked": total_booked,
                    "utilization": f"{utilization}%",
                    "status": util_status,
                })
        
        # Export to Excel
        if export:
            return self._export_excel(center.center_name, slot_bookings_data)
        
        # Pagination
        total = len(slot_bookings_data)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = slot_bookings_data[start:end]
        
        return Response({
            "success": True,
            "message": f"Slot bookings for {center.center_name}",
            "data": paginated,
            "pagination": {
                "count": total,
                "total_pages": max(1, (total + page_size - 1) // page_size),
                "current_page": page,
                "page_size": page_size,
            }
        })
    
    def _export_excel(self, center_name, slot_bookings_data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Slot Bookings"
        
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        headers = ['Slot', 'Total Booked', 'Utilization', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, sb in enumerate(slot_bookings_data, 2):
            ws.cell(row=row, column=1, value=sb['slot'])
            ws.cell(row=row, column=2, value=sb['total_booked'])
            ws.cell(row=row, column=3, value=sb['utilization'])
            ws.cell(row=row, column=4, value=sb['status'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="branch_slot_bookings_{center_name}.xlsx"'
        wb.save(response)
        return response


# ============ ADMIN REPORTS ============

class AdminCustomerReportView(APIView):
    """Get customers from all branches"""
    
    def get(self, request):
        center_id = request.query_params.get('center')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        # Get customers from all branches (or filtered by center)
        customers_qs = Customer.objects.select_related('center', 'plan').order_by('-created_at')
        if center_id:
            customers_qs = customers_qs.filter(center_id=center_id)
        if start_date:
            customers_qs = customers_qs.filter(created_at__date__gte=start_date)
        if end_date:
            customers_qs = customers_qs.filter(created_at__date__lte=end_date)
        
        customers_data = [
            {
                "id": c.id,
                "name": c.name,
                "mobile": c.mobile,
                "center": c.center.center_name if c.center else None,
                "plan": c.plan.plan_name if c.plan else None,
                "joined": str(c.created_at.date()),
                "status": c.status,
            }
            for c in customers_qs
        ]
        
        # Export to Excel
        if export:
            return self._export_excel(customers_data)
        
        # Pagination
        total = len(customers_data)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = customers_data[start:end]
        
        return Response({
            "success": True,
            "message": "Customers from all branches",
            "data": paginated,
            "pagination": {
                "count": total,
                "total_pages": max(1, (total + page_size - 1) // page_size),
                "current_page": page,
                "page_size": page_size,
            }
        })
    
    def _export_excel(self, customers_data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        headers = ['ID', 'Name', 'Mobile', 'Center', 'Plan', 'Joined', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, c in enumerate(customers_data, 2):
            ws.cell(row=row, column=1, value=c['id'])
            ws.cell(row=row, column=2, value=c['name'])
            ws.cell(row=row, column=3, value=c['mobile'])
            ws.cell(row=row, column=4, value=c['center'] or '')
            ws.cell(row=row, column=5, value=c['plan'] or '')
            ws.cell(row=row, column=6, value=c['joined'])
            ws.cell(row=row, column=7, value=c['status'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="admin_customers_all_branches.xlsx"'
        wb.save(response)
        return response


class AdminSlotBookingReportView(APIView):
    """Get slot bookings from all branches"""
    
    def get(self, request):
        center_id = request.query_params.get('center')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export = request.query_params.get('export') == 'true'
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        # Get slot bookings from all branches (or filtered by center)
        bookings_qs = SlotBooking.objects.select_related('slot', 'customer', 'center').order_by('-booking_date')
        if center_id:
            bookings_qs = bookings_qs.filter(center_id=center_id)
        if start_date:
            bookings_qs = bookings_qs.filter(booking_date__gte=start_date)
        if end_date:
            bookings_qs = bookings_qs.filter(booking_date__lte=end_date)
        
        # Group slot bookings by slot and center
        slot_bookings_data = []
        slots = Slot.objects.all().order_by('start_time')
        
        for slot in slots:
            slot_bookings = bookings_qs.filter(slot=slot)
            
            # Group by center
            centers_dict = {}
            for booking in slot_bookings:
                center_name = booking.center.center_name if booking.center else "No Center"
                if center_name not in centers_dict:
                    centers_dict[center_name] = 0
                centers_dict[center_name] += 1
            
            # Create entries for each center
            for center_name, booked_count in centers_dict.items():
                utilization = round((booked_count / slot.total_slot * 100), 1) if slot.total_slot > 0 else 0
                if utilization >= 90:
                    util_status = "high"
                elif utilization >= 70:
                    util_status = "medium"
                else:
                    util_status = "low"
                
                slot_bookings_data.append({
                    "slot": f"{slot.start_time.strftime('%I:%M %p')} - {slot.end_time.strftime('%I:%M %p')}",
                    "center": center_name,
                    "total_booked": booked_count,
                    "utilization": f"{utilization}%",
                    "status": util_status,
                })
        
        # Export to Excel
        if export:
            return self._export_excel(slot_bookings_data)
        
        # Pagination
        total = len(slot_bookings_data)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = slot_bookings_data[start:end]
        
        return Response({
            "success": True,
            "message": "Slot bookings from all branches",
            "data": paginated,
            "pagination": {
                "count": total,
                "total_pages": max(1, (total + page_size - 1) // page_size),
                "current_page": page,
                "page_size": page_size,
            }
        })
    
    def _export_excel(self, slot_bookings_data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Slot Bookings"
        
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        headers = ['Slot', 'Center', 'Total Booked', 'Utilization', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, sb in enumerate(slot_bookings_data, 2):
            ws.cell(row=row, column=1, value=sb['slot'])
            ws.cell(row=row, column=2, value=sb['center'])
            ws.cell(row=row, column=3, value=sb['total_booked'])
            ws.cell(row=row, column=4, value=sb['utilization'])
            ws.cell(row=row, column=5, value=sb['status'])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="admin_slot_bookings_all_branches.xlsx"'
        wb.save(response)
        return response

